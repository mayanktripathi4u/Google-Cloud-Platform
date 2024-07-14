# import openpyxl
from openpyxl import load_workbook
import os
import csv


source_path = r'./data/source/'
output_path = r'./data/output/'

# print(cwd = os.getcwd())
print(f"Files in given path {os.listdir(source_path)}")

for file in os.listdir(source_path):
    if not file.endswith('.xlsx'):
        continue

    file_path = os.path.join(source_path, file)

    wb = load_workbook(file_path)

    for sheetname in wb.sheetnames:
        print(f"Start conversion of Sheet {sheetname} under file {file}")

        sheet = wb[sheetname]
        csvFileName = os.path.join(output_path, f"{file[:-5]}_{sheetname}.csv")

        with open(csvFileName, 'w', newline='') as f:
            csvWriter = csv.writer(f)

            for rowNum in range(1, sheet.max_row + 1):
                rowData = []
                for colNum in range(1, sheet.max_column + 1):
                    rowData.append(sheet.cell(row=rowNum, column=colNum).value)

                csvWriter.writerow(rowData)

print("Convertion of Excel to CSV file is completed.")